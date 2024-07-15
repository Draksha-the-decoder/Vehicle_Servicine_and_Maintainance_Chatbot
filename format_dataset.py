from openpyxl import load_workbook
import csv

# Load Excel workbook
wb = load_workbook("vehicle_dataset.xlsx")
# Choose a specific sheet
sheet = wb["Brake System"]

# Find the number of rows and columns(2) in the sheet
rows = sheet.max_row
columns = sheet.max_column
print(f"Dataset has {rows} rows and {columns} columns")

# List to store all rows in the llama2 format
formatted_data = []

with open("vehicle_dataset.csv", 'w', newline='') as csvfile:
    writer = csv.writer(csvfile)
    # Iterate over rows and columns to extract data
    for i in range(2, rows+1):
        prompt = "<s>[INST] "+sheet.cell(row=i, column=1).value+" [/INST] "+sheet.cell(row=i, column=2).value+" </s>"
        writer.writerow([prompt])
print("data formatting done")

# Convert the list to a JSON string and encode it as bytes then store in the file
# with open('vehicle_dataset.json', 'wb') as f:
#     f.write(dumps(formatted_data).encode('utf-8'))

# print("data stored in vehicle_dataset.json file")